import openpyxl
import sqlite3
import datetime
from flask import flash

insert_row = []
tables = []


# connection
def create_connection(db_file):

    try:
        conn = sqlite3.connect(db_file)
        return conn
    except ConnectionError as e:
        print(e)

    return None


# insert function
def create_task(conn, task, sql):

    # review sql query
    cur = conn.cursor()
    cur.execute(sql, task)
    return cur.lastrowid


# import function from xls to punchlist table
def punchlist_import(database, importfile, log_folder):

    # create connection
    conn = create_connection(database)
    cur = conn.cursor()

    # global var
    global insert_row

    # Open Import WorkBook
    import_wb = openpyxl.load_workbook(importfile, read_only=True)
    import_ws = import_wb['punchlist']

    # other vars
    n_rows = 0
    progress = 0
    n = 2

    # open log file
    logfile = open(log_folder + str(datetime.datetime.now())+'.log', 'w')

    # cycling through excel file rows
    for row in import_ws.rows:

        insert_row = []
        for m in range(1,16):
            if import_ws.cell(row=n, column=1).value is not None:
                insert_row.append(import_ws.cell(row=n, column=m).value)


        if len(insert_row) > 0:

            n_rows += 1

            # generate datetime now for originated column date_orig
            insert_row[4] = datetime.datetime.now()

            # convert due_date excel cell value to datetime format
            due = import_ws.cell(row=n, column=6).value

            print(due)
            print(insert_row[4])

            #day = due[9:11]
            #month = due[6:8]
            #year = due[1:5]

            #insert_row[5] = datetime.datetime(int(year), int(month), int(day), 0, 0, 0, 0)
            insert_row[5] = due

            cur.execute("SELECT id from employees WHERE contactid = '" + insert_row[6] + "'")
            rows = cur.fetchall()
            insert_row[6] = rows[0][0]

            cur.execute("SELECT id from employees WHERE contactid = '" + insert_row[7] + "'")
            rows = cur.fetchall()
            insert_row[7] = rows[0][0]

            cur.execute("SELECT id from companies WHERE code = '" + insert_row[8] + "'")
            rows = cur.fetchall()
            insert_row[8] = rows[0][0]

            cur.execute("SELECT id from systems WHERE name = '" + insert_row[9] + "'")
            rows = cur.fetchall()
            try:
                insert_row[9] = rows[0][0]
            except IndexError:
                print("The system name at record " + str(n-1) + " is not present in the systems table.")
                logfile.write("The system name at record " + str(n-1) + " is not present in the systems table.\n")
                insert_row[9] = ''

            cur.execute("SELECT id from floors WHERE floor = '" + insert_row[10] + "'")
            rows = cur.fetchall()
            insert_row[10] = rows[0][0]

            cur.execute("SELECT id from phases WHERE name = '" + insert_row[11] + "'")
            rows = cur.fetchall()
            insert_row[11] = rows[0][0]

            cur.execute("SELECT id from categories WHERE name = '" + insert_row[12] + "'")
            rows = cur.fetchall()
            insert_row[12] = rows[0][0]

            cur.execute("SELECT id from buildings WHERE name = '" + str(insert_row[13]) + "'")
            rows = cur.fetchall()
            insert_row[13] = rows[0][0]

            if insert_row[14] is None:
                insert_row[14] = None
            else:
                cur.execute("SELECT id from employees WHERE contactid = '" + insert_row[14] + "'")
                rows = cur.fetchall()
                insert_row[14] = rows[0][0]

            import_row = tuple(insert_row)
            print(import_row)

            try:
                sql = ''' INSERT INTO punchlist(status,discipline,description,comments,date_orig,due_date,author_id,closure_id,supplier_id,system_id,floor_id,phase_id,cat_id,building_id,closed_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) '''
                cur.execute(sql, import_row)
                conn.commit()
                progress += 1

                print('record ' + str(progress) + ' imported successfully')
                logfile.write('record ' + str(progress) + ' imported successfully\n')

            except:
                print("An error occurred during import.")
                logfile.write("An error occurred during import.\n")

        n += 1

    import_wb.close()  # closing workbook

    # flashing summary of import operation
    flash(str(progress) + ' Items out of ' + str(n_rows) + ' imported into punchlist.')
    logfile.write(str(progress) + ' Items out of ' + str(n_rows) + ' imported into punchlist.\n')
    logfile.close() # closing logfile


# import function from xls to systems table
def systems_import(database, importfile):

    conn = create_connection(database)
    cur = conn.cursor()

    global insert_row
    # Open Import WorkBook
    import_wb = openpyxl.load_workbook(importfile, read_only=True)
    import_ws = import_wb['systems']

    n_rows = import_ws.max_row-1 # number of rows to be imported into systems
    progress = 0

    n = 2
    for row in import_ws.rows:

        progress += 1
        insert_row = []

        if import_ws.cell(row=n, column=1).value is not None:
            insert_row.append(import_ws.cell(row=n, column=1).value)

        if len(insert_row) > 0:

            import_row = tuple(insert_row)

            sql = ''' INSERT INTO systems(name) VALUES(?) '''
            cur.execute(sql,import_row)
            conn.commit()

        n += 1

    import_wb.close()
    flash(str(progress-1) + ' Items out of ' + str(n_rows) + ' imported into systems.')

#
if __name__ == "__main__":
    punchlist_import("../prologger.db","../Prologger_Import.xlsx")
