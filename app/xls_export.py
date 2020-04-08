import sys
import openpyxl
import sqlite3
import xlsxwriter
import datetime
from app.common import create_connection

insert_row = []
tables=[]


def punchlist_export(database, export_folder, images_folder, query, workbook_name=''):

    # open export file
    if workbook_name == '':
        workbook = xlsxwriter.Workbook(export_folder + 'Prologger_Export.xlsx')
    else:
        workbook = xlsxwriter.Workbook(export_folder + workbook_name + '.xlsx')

    # create a database connection
    conn = create_connection(database)
    cur = conn.cursor()

    workbook.add_worksheet('punchlist')
    worksheet = workbook.get_worksheet_by_name('punchlist')

    # get column names
    cur.execute("PRAGMA table_info(punchlist)")
    rows = cur.fetchall()

    # header format
    header_format = workbook.add_format()
    header_format.set_bold()
    header_format.set_bg_color('#808080')

    # print headers
    # add header image
    worksheet.set_row(0, 100)
    worksheet.set_column(0, 14, 24)
    worksheet.merge_range('A1:O1', '')
    worksheet.insert_image('A1', images_folder + 'sample_logo.png', {'x_offset': 15, 'y_offset': 0})

    merge_format = workbook.add_format({
    'align': 'left'})

    worksheet.merge_range('A2:O2', "Printed on: " + datetime.datetime.now().strftime("%y-%m-%d %H-%M"), merge_format)

    m = 0
    for row in rows:
        worksheet.write(2, m, row[1], header_format)
        m += 1

    # cell format
    cell_format = workbook.add_format()
    cell_format.set_text_wrap()

    # get rows for each table
    cur.execute(query)
    rows = cur.fetchall()

    # write rows into excel
    n = 3
    for row in rows:
        for m in range(len(row)):
            worksheet.write(n, m, row[m], cell_format)
        n += 1

    workbook.close()


def xls_export_extended(database, export_folder):

    # open export file
    workbook = xlsxwriter.Workbook(export_folder + 'Prologger_Export_Extended.xlsx')

    # create a database connection
    conn = create_connection(database)
    cur = conn.cursor()

    # get a list of all tables
    cur.execute("SELECT * FROM sqlite_master WHERE type='table'")
    rows = cur.fetchall()

    # add worksheet, one for each table
    for row in rows:

        # do not export sqlite_sequence and users table
        if row[1] != 'sqlite_sequence' and row[1] != 'users':
            worksheet = workbook.add_worksheet(row[1])
            tables.append(row[1])

    # fill each worksheet
    for table in tables:

        # select worksheet
        worksheet = workbook.get_worksheet_by_name(table)

        # get column names
        cur.execute("PRAGMA table_info(" + table + ")")
        rows = cur.fetchall()

        # print headers
        m = 0
        for row in rows:
            worksheet.write(0, m, row[1])
            m += 1

        # get rows for each table
        cur.execute("SELECT * FROM " + table)
        rows = cur.fetchall()

        # write rows into excel
        n = 1
        for row in rows:
            for m in range(len(row)):
                worksheet.write(n, m, row[m])
            n += 1

    workbook.close()


def punchlist_template_export(database, export_folder):

    #
    # open template file
    #

    # arrays to contains drop down lists
    buildings = []
    categories = []
    floors = []
    phases = []

    # create a database connection
    conn = create_connection(database)
    cur = conn.cursor()

    # set template file name
    workbook = xlsxwriter.Workbook(export_folder + 'Prologger_Template.xlsx')

    # setting header and cells format
    header_format = workbook.add_format()
    header_format.set_bold()
    header_format.set_bg_color('#808080')
    cell_format = workbook.add_format()
    cell_format.set_bg_color('E2EFDA')

    # execute query to build dropdown values
    cur.execute("SELECT * FROM buildings")
    rows = cur.fetchall()
    for row in rows:
        buildings.append(row[1])

    cur.execute("SELECT * FROM categories")
    rows = cur.fetchall()
    for row in rows:
        categories.append(row[1])

    cur.execute("SELECT * FROM floors")
    rows = cur.fetchall()
    for row in rows:
        floors.append(row[1])

    cur.execute("SELECT * FROM phases")
    rows = cur.fetchall()
    for row in rows:
        phases.append(row[1])

    # add worksheets to template
    worksheet = workbook.add_worksheet('punchlist')
    worksheet.set_column('A:N', 18)
    worksheet_validation = workbook.add_worksheet('validation')

    # get column names
    cur.execute("PRAGMA table_info(punchlist)")
    rows = cur.fetchall()

    # print headers
    m = 0
    for row in rows:
        if row[1] != 'id':
            worksheet.write(0, m, row[1], header_format)
            m += 1

    # create validation list for systems
    cur.execute("SELECT * FROM systems")
    rows = cur.fetchall()
    n_sys = 0
    for row in rows:
        worksheet_validation.write(n_sys, 0, row[1])
        n_sys += 1

    # create validation list for companies
    cur.execute("SELECT * FROM companies")
    rows = cur.fetchall()
    n_com = 0
    for row in rows:
        worksheet_validation.write(n_com, 1, row[1])
        n_com += 1

    # create validation list for employees
    cur.execute("SELECT * FROM employees")
    rows = cur.fetchall()
    n_emp = 0
    for row in rows:
        worksheet_validation.write(n_emp, 2, row[1])
        n_emp += 1


    # add drop down vlaues
    worksheet.data_validation('A2:A11', {'validate': 'list','source': ['0', '1']})
    worksheet.data_validation('G2:G11', {'validate': 'list','source': 'validation!$C$1:$C$' + str(n_emp)})
    worksheet.data_validation('H2:H11', {'validate': 'list','source': 'validation!$C$1:$C$' + str(n_emp)})
    worksheet.data_validation('I2:I11', {'validate': 'list','source': 'validation!$B$1:$B$' + str(n_com)})
    worksheet.data_validation('J2:J11', {'validate': 'list','source': 'validation!$A$1:$A$' + str(n_sys)})
    worksheet.data_validation('K2:K11', {'validate': 'list','source': floors})
    worksheet.data_validation('L2:L11', {'validate': 'list','source': phases})
    worksheet.data_validation('M2:M11', {'validate': 'list','source': categories})
    worksheet.data_validation('N2:N11', {'validate': 'list','source': buildings})
    worksheet.data_validation('O2:O11', {'validate': 'list','source': 'validation!$C$1:$C$' + str(n_emp)})

    # formatting cells
    for n in range(1, 11):
        for m in range(0, 14):
            worksheet.write(n, m, "", cell_format)

    # hide validation worksheet
    worksheet_validation.hide()

    # close
    workbook.close()


